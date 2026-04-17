import streamlit as st
import pandas as pd
import io
import time
from html import escape
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from utils.analysis import (
    clean_score_data, match_courses, compute_hml_breakdown,
    create_hml_chart, create_graduation_achievement_chart,
    build_indicator_course_list
)
from utils.data_loader import get_indicator_cols


UI_STEP_DELAY = 0.35


def _format_indicator_label(row):
    return f"{row['指标点ID']} {row['指标点名称']}".strip() if row['指标点名称'] else row['指标点ID']


def _brief_pause(duration=UI_STEP_DELAY):
    time.sleep(duration)


def _build_course_list_html(course_list_df):
    if course_list_df is None or course_list_df.empty:
        return ""

    style_block = """
    <style>
    .course-list-table {
        width: 100%;
        border-collapse: collapse;
        table-layout: fixed;
        margin: 0.25rem 0 0.75rem 0;
        font-size: 0.95rem;
        background: #ffffff;
        border: 1px solid #d9dee7;
    }
    .course-list-table th,
    .course-list-table td {
        border: 1px solid #d9dee7;
        padding: 10px 12px;
        vertical-align: top;
        line-height: 1.55;
    }
    .course-list-table th {
        background: #f5f7fb;
        font-weight: 600;
        text-align: center;
    }
    .course-list-table .req-cell {
        width: 18%;
        text-align: center;
        vertical-align: middle;
        font-weight: 600;
        background: #fafbfd;
    }
    .course-list-table .indicator-cell {
        width: 18%;
        text-align: center;
        vertical-align: middle;
        font-weight: 500;
    }
    .course-list-table .summary-cell {
        width: 64%;
    }
    .course-list-table .tag {
        display: inline-block;
        min-width: 1.5rem;
        padding: 1px 6px;
        margin-right: 6px;
        border-radius: 999px;
        font-size: 0.78rem;
        font-weight: 700;
        text-align: center;
        color: #ffffff;
    }
    .course-list-table .tag-h { background: #0f4c81; }
    .course-list-table .tag-m { background: #4f7cac; }
    .course-list-table .tag-l { background: #8aa9c8; }
    .course-list-table .summary-line + .summary-line {
        margin-top: 6px;
    }
    </style>
    """

    html_parts = [
        style_block,
        "<table class='course-list-table'>",
        "<thead><tr><th>毕业要求</th><th>分解指标点</th><th>关联课程清单</th></tr></thead>",
        "<tbody>",
    ]

    for _, group in course_list_df.groupby('毕业要求显示名', sort=False):
        rows = group.to_dict('records')
        rowspan = len(rows)
        for idx, row in enumerate(rows):
            indicator_label = escape(_format_indicator_label(row))
            summary_html = (
                f"<div class='summary-line'><span class='tag tag-h'>H</span>{escape(row['H课程列表'])}</div>"
                f"<div class='summary-line'><span class='tag tag-m'>M</span>{escape(row['M课程列表'])}</div>"
                f"<div class='summary-line'><span class='tag tag-l'>L</span>{escape(row['L课程列表'])}</div>"
            )

            html_parts.append("<tr>")
            if idx == 0:
                html_parts.append(
                    f"<td class='req-cell' rowspan='{rowspan}'>{escape(row['毕业要求显示名'])}</td>"
                )
            html_parts.append(f"<td class='indicator-cell'>{indicator_label}</td>")
            html_parts.append(f"<td class='summary-cell'>{summary_html}</td>")
            html_parts.append("</tr>")

    html_parts.append("</tbody></table>")
    return "".join(html_parts)


def _build_course_list_excel(course_list_df):
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        _write_course_list_sheet(writer.book, course_list_df, '指标点课程清单')

    output.seek(0)
    return output.getvalue()


def _write_course_list_sheet(workbook, course_list_df, title):
    worksheet = workbook.create_sheet(title=title)

    headers = ['毕业要求', '分解指标点', '关联课程清单']
    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color='1F2937')
        cell.fill = PatternFill(fill_type='solid', fgColor='E9EEF6')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    thin = Side(style='thin', color='D9DEE7')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    current_row = 2
    for _, group in course_list_df.groupby('毕业要求显示名', sort=False):
        start_row = current_row
        rows = group.to_dict('records')

        for row in rows:
            worksheet.cell(row=current_row, column=1, value=row['毕业要求显示名'])
            worksheet.cell(row=current_row, column=2, value=_format_indicator_label(row))
            worksheet.cell(
                row=current_row,
                column=3,
                value=f"H: {row['H课程列表']}\nM: {row['M课程列表']}\nL: {row['L课程列表']}"
            )
            current_row += 1

        end_row = current_row - 1
        if end_row > start_row:
            worksheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

    worksheet.column_dimensions['A'].width = 18
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 72

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, max_col=3):
        for cell in row:
            cell.border = border
            if cell.row == 1:
                continue
            if cell.column == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(bold=True)
            elif cell.column == 2:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    for row_idx in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_idx].height = 48


def _render_course_list_block(course_list_df, file_stub, download_key):
    if course_list_df.empty:
        st.info("当前没有可展示的指标点课程清单。")
        return

    st.markdown(_build_course_list_html(course_list_df), unsafe_allow_html=True)

    st.download_button(
        label="📥 下载课程清单（Excel合并样式）",
        data=_build_course_list_excel(course_list_df),
        file_name=f"{file_stub}_indicator_course_list.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
        key=download_key
    )

def view_analysis(df, selected_version, selected_major, program_data=None):
    """
    达成度分析视图
    """
    st.markdown(f"## 🧮 {selected_major}专业 达成度分析")
    
    st.markdown("""
    <div class="info-box">
        本模块提供两种数据输入方式和多维度分析能力。
    </div>
    """, unsafe_allow_html=True)

    
    # Check if df is valid (needed for matching)
    if df is None or df.empty:
        st.error("无法加载当前版本的矩阵数据，无法进行匹配分析。")
        return

    main_tab1, main_tab2 = st.tabs(["📤 原始数据分析", "📝 模版填报"])

    # ========================================
    # Tab 1: 原始数据上传 → 清洗 → 匹配 → 分析
    # ========================================
    with main_tab1:
        st.subheader("1. 上传原始成绩单")
        st.caption("支持教务系统导出的 Excel (含课程名称、平均分等列)")

        raw_file = st.file_uploader(
            "拖拽或点击上传", type=["xlsx", "xls"],
            key="raw_upload",
            label_visibility="collapsed"
        )

        if raw_file:
            try:
                raw_df = pd.read_excel(raw_file)

                # Step 1: Clean
                with st.status("🧹 正在清洗数据...", expanded=True) as status:
                    cleaned_df, report = clean_score_data(raw_df)

                    if 'error' in report:
                        status.update(label="❌ 清洗失败", state="error")
                        st.error(f"清洗失败: {report['error']}")
                        st.info(f"检测到的列: {raw_df.columns.tolist()}")
                        return

                    st.write(f"✅ 修正全角字符: **{report['fullwidth_fixes']}** 处")
                    st.write(f"📊 数据来源列: {report['score_source']}")
                    st.write(f"📚 有效课程数: **{report['total_courses']}** 门")
                    _brief_pause()
                    status.update(label="✅ 清洗完成", state="complete")

                # Step 2: Match
                with st.status("🔗 正在匹配矩阵...", expanded=True) as status:
                    match_log, calc_ready = match_courses(cleaned_df, df)

                    matched = len(match_log[match_log['匹配方式'] != '❌ 未匹配'])
                    unmatched = len(match_log[match_log['匹配方式'] == '❌ 未匹配'])

                    st.write(f"✅ 匹配成功: **{matched}** / {len(match_log)} 门")
                    if unmatched > 0:
                        st.write(f"❌ 未匹配: **{unmatched}** 门 (请检查课程名称差异)")
                    
                    _brief_pause()
                    status.update(label=f"✅ 匹配完成 ({matched}/{len(match_log)})", state="complete")

                # Match Details
                with st.expander("📋 查看匹配详情日志"):
                    def style_match_type(val):
                        if '✅' in str(val): return 'color: #198754'
                        if '🔶' in str(val): return 'color: #fd7e14'
                        if '❌' in str(val): return 'color: #dc3545; font-weight: bold'
                        return ''

                    st.dataframe(
                        match_log.style.map(style_match_type, subset=['匹配方式']),
                        use_container_width=True, hide_index=True
                    )
                
                st.divider()

                # Step 3: Analysis
                if calc_ready.empty:
                    st.warning("⚠️ 没有成功匹配的课程数据，无法进行达成度分析。")
                else:
                    with st.status("🧮 正在整理达成度分析视图...", expanded=True) as status:
                        st.write("正在汇总毕业要求名称与指标点课程清单...")
                        course_list_df = build_indicator_course_list(calc_ready, program_data)
                        _brief_pause()
                        st.write("正在准备图表和导出数据...")
                        _brief_pause()
                        status.update(label="✅ 分析视图已准备完成", state="complete")

                    st.success(f"🎉 准备就绪! 共 **{calc_ready['课程名称'].nunique()}** 门课程参与计算")
                    
                    # Store in session state? Not strictly necessary if we re-run, but good for persistence if we had interactions
                    # st.session_state['calc_ready_df'] = calc_ready

                    ana_tab1, ana_tab2, ana_tab3, ana_tab4 = st.tabs(
                        ["📊 毕业要求达成度", "🔬 HML 分解分析", "📚 指标点课程清单", "📥 导出报告"]
                    )

                    # --- 3a: Graduation Achievement ---
                    with ana_tab1:
                        st.subheader("毕业要求达成度")

                        # Parameters
                        with st.expander("⚙️ 计算参数 (权重 & 阈值)", expanded=False):
                            p_col1, p_col2 = st.columns([3, 1])
                            with p_col1:
                                st.caption("HML 权重设置")
                                col_w1, col_w2, col_w3 = st.columns(3)
                                raw_w_h = col_w1.number_input("H (强)", 0.0, 1.0, 0.70, 0.05, key="raw_w_h")
                                raw_w_m = col_w2.number_input("M (中)", 0.0, 1.0, 0.25, 0.05, key="raw_w_m")
                                raw_w_l = col_w3.number_input("L (弱)", 0.0, 1.0, 0.05, 0.05, key="raw_w_l")
                            with p_col2:
                                threshold = st.number_input(
                                    "达标红线", 0.0, 1.0, 0.70, 0.05,
                                    key="raw_threshold"
                                )

                        raw_weights = {'H': raw_w_h, 'M': raw_w_m, 'L': raw_w_l}

                        # Compute
                        hml_stats, hml_pivot = compute_hml_breakdown(calc_ready, weights=raw_weights)

                        # Chart
                        fig_grad = create_graduation_achievement_chart(hml_pivot, threshold)
                        st.plotly_chart(fig_grad, use_container_width=True)

                        # Table
                        st.markdown("#### 详细数据表")
                        display_pivot = hml_pivot.copy()
                        num_cols = [c for c in display_pivot.columns if c in ['H', 'M', 'L', '分解指标点达成度', '毕业要求达成度']]
                        
                        def color_warning(val):
                            if pd.isna(val) or not isinstance(val, (int, float)): return ''
                            return f'background-color: #fee2e2; color: #991b1b' if val < threshold else ''

                        st.dataframe(
                            display_pivot.style.format("{:.3f}", subset=num_cols, na_rep="-")
                                .map(color_warning, subset=num_cols),
                            use_container_width=True
                        )

                    # --- 3b: HML Breakdown ---
                    with ana_tab2:
                        st.subheader("HML 分解指标点分析")
                        
                        fig_hml = create_hml_chart(hml_stats, threshold)
                        st.plotly_chart(fig_hml, use_container_width=True)

                        # Weak points
                        weak = hml_stats[hml_stats['平均达成度'] < threshold]
                        if not weak.empty:
                            st.warning(f"⚠️ 发现 {len(weak)} 个薄弱指标点 (低于 {threshold}):")
                            st.dataframe(
                                weak[['指标点', '支撑强度', '平均达成度', '课程门数']].style.format({"平均达成度": "{:.3f}"}),
                                use_container_width=True, hide_index=True
                            )
                        else:
                            st.success("🎉 所有指标点均已达标！")

                    # --- 3c: Export ---
                    with ana_tab3:
                        st.subheader("指标点课程清单")
                        _render_course_list_block(
                            course_list_df,
                            f"achievement_analysis_{selected_major}_{selected_version}",
                            "raw_course_list_download"
                        )

                    # --- 3d: Export ---
                    with ana_tab4:
                        st.subheader("📥 下载分析报告")
                        
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            calc_ready.to_excel(writer, sheet_name='计算用长表', index=False)
                            hml_stats.to_excel(writer, sheet_name='HML详细统计', index=False)
                            hml_pivot.to_excel(writer, sheet_name='达成度透视表')
                            match_log.to_excel(writer, sheet_name='匹配日志', index=False)
                            course_list_df.to_excel(writer, sheet_name='指标点课程清单数据', index=False)
                            _write_course_list_sheet(writer.book, course_list_df, '指标点课程清单')

                        st.download_button(
                            label="📥 下载完整 Excel 报告",
                            data=output.getvalue(),
                            file_name=f"achievement_analysis_{selected_major}_{selected_version}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            type="primary"
                        )

            except Exception as e:
                st.error(f"处理文件时发生错误: {e}")
                import traceback
                st.code(traceback.format_exc())

    # ========================================
    # Tab 2: 模版填报
    # ========================================
    with main_tab2:
        st.subheader("1. 下载/上传模版")
        
        col_down, col_up = st.columns([1, 2])
        
        with col_down:
            st.markdown("#### ⬇️ 第一步: 下载模版")
            def get_template_data(version_df):
                ind_cols = get_indicator_cols(version_df)
                long_df = version_df.melt(
                    id_vars=['课程名称'],
                    value_vars=ind_cols,
                    var_name='指标点',
                    value_name='支撑强度'
                )
                long_df = long_df[long_df['支撑强度'].notna() & (long_df['支撑强度'].str.strip() != "")]
                long_df = long_df.sort_values(by=['指标点', '课程名称'])
                long_df['达成度'] = None
                return long_df

            if st.button("生成模版"):
                tpl_df = get_template_data(df)
                output_tpl = io.BytesIO()
                with pd.ExcelWriter(output_tpl, engine='openpyxl') as writer:
                    tpl_df.to_excel(writer, index=False, sheet_name='达成度填报')
                
                st.download_button(
                    label=f"📥 下载Excel模版",
                    data=output_tpl.getvalue(),
                    file_name=f"template_{selected_version}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        with col_up:
            st.markdown("#### ⬆️ 第二步: 上传已填好模版")
            uploaded_tpl = st.file_uploader("文件上传", type=["xlsx", "xls"], key="tpl_upload", label_visibility="collapsed")

        if uploaded_tpl:
            # Similar logic to original app... for brevity I'll simplify or copy logic
            # Since the original app's logic for template is quite entangled, I'll reimplement cleanly.
            try:
                input_df = pd.read_excel(uploaded_tpl)
                if '达成度' not in input_df.columns or '指标点' not in input_df.columns:
                    st.error("⚠️ 文件格式错误：缺少「达成度」或「指标点」列。")
                else:
                    st.success("文件读取成功，正在计算...")
                    # Re-use logic or implement weighted avg.
                    # Since this is "template based", we assume courses are already matched.
                    
                    # (Implementation similar to original app lines 916-1000)
                    # For modularity, I should ideally function-ize the calculation logic in utils.analysis
                    # But for now, keeping it here is fine as it's view-specific logic.
                    
                    calc_df = input_df.dropna(subset=['达成度']).copy()
                    calc_df['达成度'] = pd.to_numeric(calc_df['达成度'], errors='coerce')
                    if '支撑强度' in calc_df.columns:
                        calc_df['支撑强度'] = calc_df['支撑强度'].astype(str).str.strip().str.upper()
                    
                    # We can actually reuse compute_hml_breakdown if we adapt keys!
                    # compute_hml_breakdown expects: '指标点', '支撑强度', '达成度'
                    # It returns stats and pivot.
                    
                    # User params for template
                    with st.expander("⚙️ 计算参数"):
                        t_w_h = st.number_input("H 权重", 0.1, 1.0, 1.0, key="t_w_h")
                        t_w_m = st.number_input("M 权重", 0.1, 1.0, 0.8, key="t_w_m")
                        t_w_l = st.number_input("L 权重", 0.1, 1.0, 0.6, key="t_w_l")
                        t_thresh = st.slider("阈值", 0.0, 1.0, 0.7, key="t_thresh")
                    
                    w_map = {'H': t_w_h, 'M': t_w_m, 'L': t_w_l}

                    with st.status("🧮 正在计算模板分析结果...", expanded=True) as status:
                        st.write("正在计算 HML 达成度与毕业要求达成度...")
                        hml_stats, hml_pivot = compute_hml_breakdown(calc_df, weights=w_map)
                        _brief_pause()
                        st.write("正在整理指标点课程清单...")
                        course_list_df = build_indicator_course_list(calc_df, program_data)
                        _brief_pause()
                        status.update(label="✅ 模板分析完成", state="complete")
                    
                    st.markdown("#### 📊 分析结果")
                    fig_grad_tpl = create_graduation_achievement_chart(hml_pivot, t_thresh)
                    st.plotly_chart(fig_grad_tpl, use_container_width=True)
                    
                    st.dataframe(hml_pivot.style.format("{:.3f}", na_rep="-"), use_container_width=True)

                    if not course_list_df.empty:
                        st.markdown("#### 📚 指标点课程清单")
                        _render_course_list_block(
                            course_list_df,
                            f"template_analysis_{selected_major}_{selected_version}",
                            "template_course_list_download"
                        )

            except Exception as e:
                st.error(f"Error: {e}")
